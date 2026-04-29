from __future__ import annotations

import os
import csv
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
EXAMPLE = ROOT / "examples" / "meta_sample.csv"


class TestSmoke(unittest.TestCase):
    def _run(self, args: list[str]) -> subprocess.CompletedProcess[str]:
        env = os.environ.copy()
        env["PYTHONPATH"] = str(SRC)
        return subprocess.run(
            [sys.executable, "-m", "adswk", *args],
            cwd=str(ROOT),
            env=env,
            text=True,
            capture_output=True,
        )

    def test_smoke_generates_outputs(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            out_dir = Path(td) / "outputs"
            p = self._run(
                [
                    "report",
                    "--input",
                    str(EXAMPLE),
                    "--output",
                    str(out_dir),
                    "--template",
                    "meta_ads",
                    "--level",
                    "campaign",
                ]
            )
            self.assertEqual(p.returncode, 0, msg=f"stdout:\n{p.stdout}\n\nstderr:\n{p.stderr}")
            html_path = out_dir / "weekly_report.html"
            xlsx_path = out_dir / "clean.xlsx"
            self.assertTrue(html_path.exists())
            self.assertTrue(xlsx_path.exists())

            html = html_path.read_text(encoding="utf-8")
            self.assertIn("KPI", html)
            self.assertIn("Top 10", html)
            self.assertIn("Anomaly", html)
            self.assertGreaterEqual(html.count("data:image/png;base64,"), 2)

            wb = openpyxl.load_workbook(xlsx_path)
            try:
                self.assertIn("clean_data", wb.sheetnames)
                self.assertIn("summary", wb.sheetnames)
            finally:
                wb.close()

    def test_missing_required_field_errors(self) -> None:
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tmp_csv = Path(td) / "missing_impressions.csv"
            with EXAMPLE.open("r", encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
            header = rows[0]
            idx = header.index("Impressions")
            header = [c for i, c in enumerate(header) if i != idx]
            out_rows = [header]
            for row in rows[1:]:
                out_rows.append([c for i, c in enumerate(row) if i != idx])
            with tmp_csv.open("w", encoding="utf-8", newline="") as f:
                w = csv.writer(f)
                w.writerows(out_rows)

            out_dir = Path(td) / "outputs"
            p = self._run(
                [
                    "report",
                    "--input",
                    str(tmp_csv),
                    "--output",
                    str(out_dir),
                    "--template",
                    "meta_ads",
                ]
            )
            self.assertNotEqual(p.returncode, 0)
            self.assertIn("Missing required fields", p.stderr)
            self.assertIn("请在 Meta Ads Reporting 导出时包含", p.stderr)

    def test_kpi_formula_alignment(self) -> None:
        env = os.environ.copy()
        env["PYTHONPATH"] = str(SRC)
        code = (
            "import pandas as pd\n"
            "from adswk.kpi import add_kpi_columns\n"
            "df=pd.DataFrame({'spend':[10.0,10.0],'impressions':[100.0,0.0],'link_clicks':[5.0,0.0]})\n"
            "out=add_kpi_columns(df)\n"
            "assert abs(out.loc[0,'ctr']-0.05)<1e-9\n"
            "assert abs(out.loc[0,'cpc']-2.0)<1e-9\n"
            "assert abs(out.loc[0,'cpm']-100.0)<1e-9\n"
            "assert out.loc[1,'ctr']==0.0\n"
            "assert str(out.loc[1,'cpc'])=='nan'\n"
            "assert str(out.loc[1,'cpm'])=='nan'\n"
            "print('ok')\n"
        )
        p = subprocess.run([sys.executable, "-c", code], env=env, text=True, capture_output=True)
        self.assertEqual(p.returncode, 0, msg=p.stderr)


if __name__ == "__main__":
    unittest.main()
